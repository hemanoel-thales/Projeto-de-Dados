import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime


diretorio = "C:/Users/PC Gamer/Desktop/Python/Extração PDF/Documento PDF" #Define o diretório dos arquivos PDF 
arquivos = os.listdir(diretorio) #Lista todos os arquivos existentes no diretório
qtd_arquivos = len(arquivos) #Conta o total de arquivos no diretório

if qtd_arquivos == 0: #Verifica se há arquivos no diretório
    raise Exception("Não foram encontrados arquivos PDF") #Caso não tenha arquivos, encerra o código e informa uma mensagem de erro

wb = Workbook() #Define a variável para gerar um WorkBook Excel
ws = wb.active #Define a variavel para o WorkSheet aberta no WorkBook
ws.title = "Lista de Fornecedores" #Renomeia o WorkSheet

#Define o nome dos cabeçalhos da WorkSheet
ws["A1"] = "Fornecedor #"
ws["B1"] = "Instagram"
ws["C1"] = "Link Insta"
ws["D1"] = "Descrição"
ws["E1"] = "Pedidos"
ws["F1"] = "Status"

ultima_linha = 2 #Define uma variável para controlar as linhas em branco

for arquivo in arquivos: #Inicia o loop dentro dos arquivos enconrados no diretório
    with pdfplumber.open(diretorio + "/" + arquivo) as pdf: #abre o primeiro arquivo pdf da pasta

        total_paginas = len(pdf.pages) #Define o total de páginas do arquivo pdf aberto

        for pagina in range(6,total_paginas): #inicia um loop por todas as paginas do PDF, iniciando pela página 6
            primeira_pagina = pdf.pages[pagina] #Define o indice da pagina atual do loop
            texto_pagina = primeira_pagina.extract_text().upper() #Captura todo o texto da pagina atual e transforma tudo em maiusculas

            re_fornecedor = r'FORNECEDOR #(\d{3})' #Pesquisa o que vem após o padrão 'FORNECEDOR #' + 3 dígitos
            re_instagram1 = r'INSTAGRAM @(\w.+)' #Pesquisa o que vem após o padrão 'INSTAGRAM @', pegando toda a string com palavras, numeros e pontos
            re_instagram2 = r'INSTAGRAM: @(\w.+)' #Pesquisa o que vem após o padrão Instragram, mas adicionando ': ' ao final
            re_descricao = r'TRABALHA COM (.+)' #Pesquisa tudo que vier depois de 'TRABALHA COM '
            re_pedidos1 = r'MÍNIMO (.+)' #Pesquisa tudo que vier depois de 'MÍNIMO ' com acento
            re_pedidos2 = r'MINIMO (.+)' #Pesquisa tudo que vier depois de 'MINIMO ' sem acento

            #Pesquisa no texto da página utilizando o parâmetro passado anteriormente
            match_fornecedor = re.search(re_fornecedor, texto_pagina) 
            match_instagram1 = re.search(re_instagram1, texto_pagina)
            match_instagram2 = re.search(re_instagram2, texto_pagina)
            match_descricao = re.search(re_descricao, texto_pagina)
            match_pedidos1 = re.search(re_pedidos1, texto_pagina)
            match_pedidos2 = re.search(re_pedidos2, texto_pagina)

            if match_fornecedor: # Verifica se foi encontrado alguma coisa na página conforme o padrão informado no Regex
                fornecedor = match_fornecedor.group(1) # Captura o resultado da consulta no texto
                ws[f"A{ultima_linha}"] = fornecedor # Adiciona o valor no WorkSheet
            else:
                ws[f"A{ultima_linha}"] = "Não encontrado" # Caso não encontre, adiciona uma mensagem

            if match_instagram1:
                instagram = match_instagram1.group(1)
                ws[f"B{ultima_linha}"] = instagram
                ws[f"C{ultima_linha}"] = f"instagram.com/{instagram}/" # Cria a URL do usuario do Instagram
            elif match_instagram2:
                instagram = match_instagram2.group(1)
                ws[f"B{ultima_linha}"] = instagram
                ws[f"C{ultima_linha}"] = f"instagram.com/{instagram}/"
            else:
                ws[f"B{ultima_linha}"] = "Não encontrado"
                ws[f"C{ultima_linha}"] = "Não encontrado"

            if match_descricao:
                descricao = match_descricao.group(1)
                ws[f"D{ultima_linha}"] = descricao
            else:
                ws[f"D{ultima_linha}"] = "Não encontrado"
            
            if match_pedidos1:
                pedido = match_pedidos1.group(1)
                ws[f"E{ultima_linha}"] = pedido
            elif match_pedidos2:
                pedido = match_pedidos2.group(1)
                ws[f"E{ultima_linha}"] = pedido
            else:
                ws[f"E{ultima_linha}"] = "Não encontrado"
            
            ws[f"F{ultima_linha}"] = "Completo"

            ultima_linha += 1

agora_completo = str(datetime.now()).replace(":","-") # Captura a data e hora atual
agora = agora_completo[:19] # Retira o milisegundos da hora

wb.save(f"Fornecedores - {agora}.xlsx") # Salva o arquivo Excel com todas as informações, considerando a hora e datas atuais




